#!/opt/anaconda3/bin/python3
"""Sync processing_log.jsonl's video_done events into report.xlsx.

Incremental, not a rebuild: reads whatever rows are already in report.xlsx
(keyed by the video's filename, not its full path -- the full path includes
transient extraction folders that vary by machine/run, while the filename
itself, camera + timestamp, is what actually identifies a unique recording),
and only appends rows for video_done events not already present -- existing
rows are left untouched. Safe to run repeatedly, standalone or as the last
step of automate.py, since it never recreates the file.

Usage:
    ./export_report.py
"""
import json
import re
from datetime import date as date_cls, time as time_cls
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent
LOG_PATH = ROOT / "processing_log 1.jsonl"
EXCEL_PATH = ROOT / "report.xlsx"
SHEET_NAME = "Videos"

# Same two filename shapes merge_active.py recognizes for cut clips, applied
# here to the raw video's own filename (its stem, with no "_segNNN_..." suffix).
DATE_TIME_RE = re.compile(
    r"Camera_(?P<date1>\d{4}-\d{2}-\d{2})_(?P<time1>\d{2}-\d{2}-\d{2})"
    r"|cam-(?P<date2>\d{8})-(?P<time2>\d{6})"
)

DATA_COLUMNS = [
    "video", "duration_sec", "active_segments", "idle_segments",
    "active_time_sec", "idle_time_sec", "corrupted_segments",
]
COLUMNS = ["date", "time"] + DATA_COLUMNS


def parse_video_date_time(video_path: str):
    """Pulls the recording date/time out of the video's own filename, as
    real date/time objects (not strings) so Excel treats them as sortable,
    filterable columns rather than text. Returns (None, None) if the
    filename doesn't match either known naming convention."""
    m = DATE_TIME_RE.search(Path(video_path).stem)
    if not m:
        return None, None
    if m.group("date1"):
        y, mo, d = (int(x) for x in m.group("date1").split("-"))
        h, mi, s = (int(x) for x in m.group("time1").split("-"))
    else:
        date2, time2 = m.group("date2"), m.group("time2")
        y, mo, d = int(date2[0:4]), int(date2[4:6]), int(date2[6:8])
        h, mi, s = int(time2[0:2]), int(time2[2:4]), int(time2[4:6])
    return date_cls(y, mo, d), time_cls(h, mi, s)


def load_or_create_sheet():
    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.create_sheet(SHEET_NAME)
        if ws.max_row == 1 and ws.cell(1, 1).value is None:
            ws.append(COLUMNS)
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(COLUMNS)
    return wb, ws


def existing_videos(ws):
    header = [c.value for c in ws[1]]
    col = header.index("video")
    videos = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[col]:
            videos.add(row[col])
    return videos


def video_done_events():
    if not LOG_PATH.exists():
        return
    with LOG_PATH.open() as f:
        for line in f:
            try:
                d = json.loads(line)
            except json.JSONDecodeError:
                continue
            if d.get("event") == "video_done":
                yield d


def main():
    wb, ws = load_or_create_sheet()
    done = existing_videos(ws)

    added = 0
    for event in video_done_events():
        video_name = Path(event["video"]).name
        if video_name in done:
            continue
        video_date, video_time = parse_video_date_time(event["video"])
        row = [video_date, video_time, video_name] + [event.get(col) for col in DATA_COLUMNS[1:]]
        ws.append(row)
        done.add(video_name)
        added += 1

    if added:
        wb.save(EXCEL_PATH)
        print(f"added {added} new row(s) to {EXCEL_PATH}")
    else:
        print(f"no new video_done events since last sync ({EXCEL_PATH} unchanged)")


if __name__ == "__main__":
    main()
